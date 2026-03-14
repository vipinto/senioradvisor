from fastapi import APIRouter, Request, HTTPException, UploadFile, File
from pydantic import BaseModel
from typing import List, Optional
from datetime import datetime, timezone
import uuid, random, string, io
from passlib.hash import bcrypt

from database import db
from auth import get_current_user, require_admin
from routes.notification_routes import create_notification

router = APIRouter(prefix="/admin")


class PlanCreateUpdate(BaseModel):
    name: str
    duration_months: int
    price_clp: int
    features: List[str] = []
    popular: bool = False


# ============= PROVIDER MANAGEMENT =============

@router.get("/providers/pending")
async def get_pending_providers(request: Request):
    """Get providers awaiting approval"""
    user = await get_current_user(request, db)
    await require_admin(user)
    providers = await db.providers.find({"approved": False}, {"_id": 0}).to_list(100)
    return providers


@router.get("/providers/all")
async def get_all_providers(request: Request):
    """Get all providers for admin"""
    user = await get_current_user(request, db)
    await require_admin(user)
    providers = await db.providers.find({}, {"_id": 0}).sort("created_at", -1).to_list(200)
    return providers


@router.post("/providers/{provider_id}/approve")
async def approve_provider(provider_id: str, request: Request):
    """Approve provider"""
    user = await get_current_user(request, db)
    await require_admin(user)

    result = await db.providers.update_one(
        {"provider_id": provider_id},
        {"$set": {"approved": True, "approved_at": datetime.now(timezone.utc)}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")

    provider = await db.providers.find_one({"provider_id": provider_id})
    if provider:
        await create_notification(
            user_id=provider["user_id"],
            title="¡Tu perfil fue aprobado!",
            message="Tu perfil de proveedor ha sido aprobado. Ya apareces en las búsquedas.",
            notification_type="provider_approved"
        )
    return {"message": "Proveedor aprobado"}


@router.post("/providers/{provider_id}/reject")
async def reject_provider(provider_id: str, request: Request):
    """Reject provider"""
    user = await get_current_user(request, db)
    await require_admin(user)

    body = await request.json()
    reason = body.get("reason", "No cumple con los requisitos")

    provider = await db.providers.find_one({"provider_id": provider_id})
    if not provider:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")

    await db.providers.delete_one({"provider_id": provider_id})
    await db.users.update_one(
        {"user_id": provider["user_id"]},
        {"$set": {"role": "user"}}
    )

    await create_notification(
        user_id=provider["user_id"],
        title="Perfil rechazado",
        message=f"Tu perfil de proveedor fue rechazado. Razón: {reason}",
        notification_type="provider_rejected"
    )
    return {"message": "Proveedor rechazado"}


@router.post("/providers/{provider_id}/verify")
async def verify_provider(provider_id: str, request: Request):
    """Mark provider as verified"""
    user = await get_current_user(request, db)
    await require_admin(user)

    result = await db.providers.update_one(
        {"provider_id": provider_id},
        {"$set": {"verified": True, "verified_at": datetime.now(timezone.utc)}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")

    provider = await db.providers.find_one({"provider_id": provider_id})
    if provider:
        await create_notification(
            user_id=provider["user_id"],
            title="¡Cuenta verificada!",
            message="Tu cuenta ha sido verificada. Ahora tienes el badge de proveedor verificado.",
            notification_type="provider_verified"
        )
    return {"message": "Proveedor verificado"}


@router.post("/providers/{provider_id}/unverify")
async def unverify_provider(provider_id: str, request: Request):
    """Remove verified badge"""
    user = await get_current_user(request, db)
    await require_admin(user)

    result = await db.providers.update_one(
        {"provider_id": provider_id},
        {"$set": {"verified": False, "verified_at": None}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")
    return {"message": "Verificación removida"}


# ============= STATS & METRICS =============

@router.get("/stats")
async def get_admin_stats(request: Request):
    """Get admin dashboard stats"""
    user = await get_current_user(request, db)
    await require_admin(user)

    total_users = await db.users.count_documents({})
    total_providers = await db.providers.count_documents({"approved": True})
    pending_providers = await db.providers.count_documents({"approved": False})
    verified_providers = await db.providers.count_documents({"verified": True})
    active_subscriptions = await db.subscriptions.count_documents({"status": "active"})
    total_reviews = await db.reviews.count_documents({})

    return {
        "total_users": total_users,
        "total_providers": total_providers,
        "pending_providers": pending_providers,
        "verified_providers": verified_providers,
        "active_subscriptions": active_subscriptions,
        "total_reviews": total_reviews
    }


@router.get("/metrics")
async def get_admin_metrics(request: Request):
    """Get time-series metrics for admin dashboard charts"""
    user = await get_current_user(request, db)
    await require_admin(user)

    months = []
    now = datetime.now(timezone.utc)
    for i in range(5, -1, -1):
        month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        if i > 0:
            m = month_start.month - i
            y = month_start.year
            while m <= 0:
                m += 12
                y -= 1
            month_start = month_start.replace(year=y, month=m)

        next_month = month_start.month + 1
        next_year = month_start.year
        if next_month > 12:
            next_month = 1
            next_year += 1
        month_end = month_start.replace(year=next_year, month=next_month)

        users_count = await db.users.count_documents({
            "created_at": {"$gte": month_start, "$lt": month_end}
        })
        providers_count = await db.providers.count_documents({
            "created_at": {"$gte": month_start, "$lt": month_end}
        })
        subs_count = await db.subscriptions.count_documents({
            "start_date": {"$gte": month_start, "$lt": month_end}
        })
        reviews_count = await db.reviews.count_documents({
            "created_at": {"$gte": month_start, "$lt": month_end}
        })

        month_names = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
        months.append({
            "month": month_names[month_start.month - 1],
            "users": users_count,
            "providers": providers_count,
            "subscriptions": subs_count,
            "reviews": reviews_count
        })

    return months


# ============= PLAN MANAGEMENT =============

@router.get("/plans")
async def get_all_plans(request: Request):
    """Get all plans for admin"""
    user = await get_current_user(request, db)
    await require_admin(user)
    plans = await db.subscription_plans.find({}, {"_id": 0}).sort("price_clp", 1).to_list(50)
    return plans


@router.post("/plans")
async def create_plan(data: PlanCreateUpdate, request: Request):
    """Create a new subscription plan"""
    user = await get_current_user(request, db)
    await require_admin(user)

    plan_id = f"plan_{uuid.uuid4().hex[:8]}"
    plan = {
        "plan_id": plan_id,
        **data.model_dump(),
        "active": True,
        "created_at": datetime.now(timezone.utc)
    }
    await db.subscription_plans.insert_one(plan)
    plan.pop("_id", None)
    return plan


@router.put("/plans/{plan_id}")
async def update_plan(plan_id: str, data: PlanCreateUpdate, request: Request):
    """Update a subscription plan"""
    user = await get_current_user(request, db)
    await require_admin(user)

    result = await db.subscription_plans.update_one(
        {"plan_id": plan_id},
        {"$set": {**data.model_dump(), "updated_at": datetime.now(timezone.utc)}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Plan no encontrado")

    updated = await db.subscription_plans.find_one({"plan_id": plan_id}, {"_id": 0})
    return updated


@router.post("/plans/{plan_id}/toggle")
async def toggle_plan(plan_id: str, request: Request):
    """Activate/deactivate a plan"""
    user = await get_current_user(request, db)
    await require_admin(user)

    plan = await db.subscription_plans.find_one({"plan_id": plan_id})
    if not plan:
        raise HTTPException(status_code=404, detail="Plan no encontrado")

    new_active = not plan.get("active", True)
    await db.subscription_plans.update_one(
        {"plan_id": plan_id},
        {"$set": {"active": new_active}}
    )
    return {"message": f"Plan {'activado' if new_active else 'desactivado'}", "active": new_active}


@router.delete("/plans/{plan_id}")
async def delete_plan(plan_id: str, request: Request):
    """Delete a subscription plan"""
    user = await get_current_user(request, db)
    await require_admin(user)

    result = await db.subscription_plans.delete_one({"plan_id": plan_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Plan no encontrado")
    return {"message": "Plan eliminado"}


# ============= MAKE ADMIN =============

@router.post("/make-admin")
async def make_admin(request: Request):
    """Make a user admin by email"""
    body = await request.json()
    email = body.get("email")
    if not email:
        raise HTTPException(status_code=400, detail="Email requerido")

    admin_count = await db.users.count_documents({"role": "admin"})
    if admin_count > 0:
        user = await get_current_user(request, db)
        await require_admin(user)

    result = await db.users.update_one(
        {"email": email},
        {"$set": {"role": "admin"}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    return {"message": f"Usuario {email} ahora es admin"}


# ============= SOS CONFIGURATION =============

@router.get("/sos")
async def get_sos_config(request: Request):
    """Get SOS configuration"""
    user = await get_current_user(request, db)
    await require_admin(user)
    config = await db.sos_config.find_one({}, {"_id": 0})
    if not config:
        return {"active": False, "phone": "", "schedule": "", "vet_name": "", "start_hour": 8, "end_hour": 20}
    return config


@router.put("/sos")
async def update_sos_config(request: Request):
    """Update SOS configuration"""
    user = await get_current_user(request, db)
    await require_admin(user)
    data = await request.json()

    allowed = ['phone', 'schedule', 'vet_name', 'active', 'start_hour', 'end_hour']
    update_data = {k: v for k, v in data.items() if k in allowed}
    update_data["updated_at"] = datetime.now(timezone.utc)

    await db.sos_config.update_one(
        {},
        {"$set": update_data},
        upsert=True
    )
    config = await db.sos_config.find_one({}, {"_id": 0})
    return config


# --- Create Residencia ---

class ResidenciaCreate(BaseModel):
    business_name: str
    email: str
    password: Optional[str] = None
    phone: Optional[str] = ""
    whatsapp: Optional[str] = ""
    address: Optional[str] = ""
    comuna: Optional[str] = ""
    description: Optional[str] = ""
    service_type: Optional[str] = "residencias"
    price_from: Optional[int] = 0

def generate_password(length=10):
    chars = string.ascii_letters + string.digits
    return ''.join(random.choice(chars) for _ in range(length))

@router.post("/residencias/create")
async def create_residencia(data: ResidenciaCreate, request: Request):
    user = await get_current_user(request, db)
    await require_admin(user)
    
    existing = await db.users.find_one({"email": data.email})
    if existing:
        raise HTTPException(status_code=400, detail=f"El email {data.email} ya está registrado")
    
    password = data.password or generate_password()
    user_id = str(uuid.uuid4())
    provider_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    
    user = {
        "user_id": user_id,
        "email": data.email,
        "name": data.business_name,
        "role": "provider",
        "hashed_password": bcrypt.hash(password),
        "created_at": now.isoformat(),
        "active": True,
    }
    await db.users.insert_one(user)
    
    provider = {
        "provider_id": provider_id,
        "user_id": user_id,
        "business_name": data.business_name,
        "phone": data.phone or "",
        "whatsapp": data.whatsapp or "",
        "address": data.address or "",
        "comuna": data.comuna or "",
        "description": data.description or "",
        "services": [{"service_type": data.service_type or "residencias", "price_from": data.price_from or 0, "description": ""}],
        "photos": [],
        "gallery": [],
        "amenities": [],
        "social_links": {},
        "personal_info": {},
        "rating": 0,
        "total_reviews": 0,
        "approved": True,
        "verified": False,
        "latitude": 0,
        "longitude": 0,
        "coverage_zone": "10",
        "created_at": now,
        "approved_at": now,
    }
    await db.providers.insert_one(provider)
    
    return {
        "provider_id": provider_id,
        "user_id": user_id,
        "business_name": data.business_name,
        "email": data.email,
        "password": password,
        "status": "created"
    }

class BulkResidenciaItem(BaseModel):
    business_name: str
    email: str
    phone: Optional[str] = ""
    whatsapp: Optional[str] = ""
    address: Optional[str] = ""
    comuna: Optional[str] = ""
    description: Optional[str] = ""
    service_type: Optional[str] = "residencias"
    price_from: Optional[int] = 0

class BulkResidenciaCreate(BaseModel):
    residencias: List[BulkResidenciaItem]

@router.post("/residencias/bulk-create")
async def bulk_create_residencias(data: BulkResidenciaCreate, request: Request):
    user = await get_current_user(request, db)
    await require_admin(user)
    
    results = []
    now = datetime.now(timezone.utc)
    
    for item in data.residencias:
        existing = await db.users.find_one({"email": item.email})
        if existing:
            results.append({"business_name": item.business_name, "email": item.email, "status": "error", "detail": "Email ya registrado"})
            continue
        
        password = generate_password()
        user_id = str(uuid.uuid4())
        provider_id = str(uuid.uuid4())
        
        user = {
            "user_id": user_id,
            "email": item.email,
            "name": item.business_name,
            "role": "provider",
            "hashed_password": bcrypt.hash(password),
            "created_at": now.isoformat(),
            "active": True,
        }
        await db.users.insert_one(user)
        
        provider = {
            "provider_id": provider_id,
            "user_id": user_id,
            "business_name": item.business_name,
            "phone": item.phone or "",
            "whatsapp": item.whatsapp or "",
            "address": item.address or "",
            "comuna": item.comuna or "",
            "description": item.description or "",
            "services": [{"service_type": item.service_type or "residencias", "price_from": item.price_from or 0, "description": ""}],
            "photos": [],
            "gallery": [],
            "amenities": [],
            "social_links": {},
            "personal_info": {},
            "rating": 0,
            "total_reviews": 0,
            "approved": True,
            "verified": False,
            "latitude": 0,
            "longitude": 0,
            "coverage_zone": "10",
            "created_at": now,
            "approved_at": now,
        }
        await db.providers.insert_one(provider)
        
        results.append({
            "business_name": item.business_name,
            "email": item.email,
            "password": password,
            "provider_id": provider_id,
            "status": "created"
        })
    
    created = len([r for r in results if r["status"] == "created"])
    errors = len([r for r in results if r["status"] == "error"])
    return {"total": len(results), "created": created, "errors": errors, "results": results}


@router.post("/residencias/upload-excel")
async def upload_excel_residencias(request: Request, file: UploadFile = File(...)):
    user = await get_current_user(request, db)
    await require_admin(user)
    
    import openpyxl
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    
    headers = [str(cell.value or "").strip().lower() for cell in ws[1]]
    
    # Map common column names
    col_map = {}
    for i, h in enumerate(headers):
        if h in ("nombre", "business_name", "residencia", "nombre residencia", "nombre_residencia"):
            col_map["business_name"] = i
        elif h in ("email", "correo", "correo electrónico", "correo electronico", "mail"):
            col_map["email"] = i
        elif h in ("telefono", "teléfono", "phone", "fono"):
            col_map["phone"] = i
        elif h in ("whatsapp", "wsp"):
            col_map["whatsapp"] = i
        elif h in ("direccion", "dirección", "address"):
            col_map["address"] = i
        elif h in ("comuna", "ciudad"):
            col_map["comuna"] = i
        elif h in ("descripcion", "descripción", "description"):
            col_map["description"] = i
        elif h in ("tipo", "tipo servicio", "tipo_servicio", "service_type", "categoria"):
            col_map["service_type"] = i
        elif h in ("precio", "price", "precio_desde", "price_from"):
            col_map["price_from"] = i
    
    if "business_name" not in col_map or "email" not in col_map:
        raise HTTPException(status_code=400, detail="El archivo debe tener al menos columnas 'nombre' y 'email'")
    
    results = []
    now = datetime.now(timezone.utc)
    
    for row in ws.iter_rows(min_row=2, values_only=False):
        values = [str(cell.value or "").strip() for cell in row]
        bname = values[col_map["business_name"]]
        email = values[col_map["email"]]
        
        if not bname or not email:
            continue
        
        existing = await db.users.find_one({"email": email})
        if existing:
            results.append({"business_name": bname, "email": email, "status": "error", "detail": "Email ya registrado"})
            continue
        
        password = generate_password()
        user_id = str(uuid.uuid4())
        provider_id = str(uuid.uuid4())
        
        phone = values[col_map["phone"]] if "phone" in col_map else ""
        whatsapp = values[col_map["whatsapp"]] if "whatsapp" in col_map else phone
        address = values[col_map["address"]] if "address" in col_map else ""
        comuna = values[col_map["comuna"]] if "comuna" in col_map else ""
        description = values[col_map["description"]] if "description" in col_map else ""
        service_type = values[col_map["service_type"]] if "service_type" in col_map else "residencias"
        
        price_from = 0
        if "price_from" in col_map:
            try:
                price_from = int(float(values[col_map["price_from"]].replace(".", "").replace("$", "").replace(",", "")))
            except:
                pass
        
        # Normalize service type
        st_lower = service_type.lower()
        if "domicilio" in st_lower:
            service_type = "cuidado_domicilio"
        elif "mental" in st_lower or "psico" in st_lower:
            service_type = "salud_mental"
        else:
            service_type = "residencias"
        
        user = {
            "user_id": user_id,
            "email": email,
            "name": bname,
            "role": "provider",
            "hashed_password": bcrypt.hash(password),
            "created_at": now.isoformat(),
            "active": True,
        }
        await db.users.insert_one(user)
        
        provider = {
            "provider_id": provider_id,
            "user_id": user_id,
            "business_name": bname,
            "phone": phone,
            "whatsapp": whatsapp or phone,
            "address": address,
            "comuna": comuna,
            "description": description,
            "services": [{"service_type": service_type, "price_from": price_from, "description": ""}],
            "photos": [],
            "gallery": [],
            "amenities": [],
            "social_links": {},
            "personal_info": {},
            "rating": 0,
            "total_reviews": 0,
            "approved": True,
            "verified": False,
            "latitude": 0,
            "longitude": 0,
            "coverage_zone": "10",
            "created_at": now,
            "approved_at": now,
        }
        await db.providers.insert_one(provider)
        
        results.append({
            "business_name": bname,
            "email": email,
            "password": password,
            "provider_id": provider_id,
            "status": "created"
        })
    
    created = len([r for r in results if r["status"] == "created"])
    errors = len([r for r in results if r["status"] == "error"])
    return {"total": len(results), "created": created, "errors": errors, "results": results}
